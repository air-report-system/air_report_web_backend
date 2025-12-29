"""
月度账表 AI WebSocket 消费者

协议（与前端约定）：
- subscribe_report: { report_id }
- ai_calculate: { report_id, question, sheet? }
推送：
- progress: { stage, message }
- final: { label, value, write_cell, preview }
- error: { error }
"""
import json
import logging
from typing import Any, Dict, Optional

from channels.generic.websocket import AsyncWebsocketConsumer
from channels.db import database_sync_to_async
from django.contrib.auth.models import AnonymousUser

logger = logging.getLogger(__name__)


class MonthlyReportAIConsumer(AsyncWebsocketConsumer):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.user = None
        self.report_id: Optional[int] = None
        self.group_name: Optional[str] = None

    async def connect(self):
        try:
            self.user = self.scope["user"]
            if isinstance(self.user, AnonymousUser):
                logger.warning("未认证用户尝试连接Monthly AI WebSocket")
                await self.close(code=4001)
                return

            await self.accept()
            await self._send('connection_established', {
                'user_id': self.user.id,
                'username': getattr(self.user, 'username', ''),
                'message': 'WebSocket连接已建立'
            })
        except Exception as e:
            logger.error(f"Monthly WS connect error: {e}")
            await self.close(code=4000)

    async def disconnect(self, close_code):
        try:
            if self.group_name:
                await self.channel_layer.group_discard(self.group_name, self.channel_name)
        except Exception as e:
            logger.error(f"Monthly WS disconnect error: {e}")

    async def receive(self, text_data):
        try:
            payload = json.loads(text_data)
            message_type = payload.get('type')
            data = payload.get('data', {}) or {}

            if message_type == 'ping':
                await self._send('pong', {})
            elif message_type == 'subscribe_report':
                await self._handle_subscribe_report(data)
            elif message_type == 'ai_calculate':
                await self._handle_ai_calculate(data)
            else:
                await self._send('error', {'error': f'未知消息类型: {message_type}'})
        except json.JSONDecodeError:
            await self._send('error', {'error': '无效的JSON数据'})
        except Exception as e:
            logger.error(f"Monthly WS receive error: {e}")
            await self._send('error', {'error': f'处理消息时出错: {str(e)}'})

    async def _handle_subscribe_report(self, data: Dict[str, Any]):
        report_id = data.get('report_id')
        if not report_id:
            await self._send('error', {'error': 'report_id is required'})
            return

        try:
            report_id = int(report_id)
        except (ValueError, TypeError):
            await self._send('error', {'error': 'report_id must be int'})
            return

        report = await self._get_report(report_id)
        if not report:
            await self._send('error', {'error': f'报表 {report_id} 不存在或无权限访问'})
            return

        # 切换订阅
        if self.group_name:
            await self.channel_layer.group_discard(self.group_name, self.channel_name)

        self.report_id = report_id
        self.group_name = f"monthly_report_{report_id}"
        await self.channel_layer.group_add(self.group_name, self.channel_name)

        await self._send('subscribed', {'report_id': report_id})

    async def _handle_ai_calculate(self, data: Dict[str, Any]):
        report_id = data.get('report_id') or self.report_id
        question = (data.get('question') or '').strip()
        sheet = (data.get('sheet') or '账单明细').strip() or '账单明细'

        if not report_id:
            await self._send('error', {'error': 'report_id is required'})
            return
        if not question:
            await self._send('error', {'error': 'question is required'})
            return

        try:
            report_id = int(report_id)
        except (ValueError, TypeError):
            await self._send('error', {'error': 'report_id must be int'})
            return

        report = await self._get_report(report_id)
        if not report:
            await self._send('error', {'error': f'报表 {report_id} 不存在或无权限访问'})
            return

        await self._send('progress', {'stage': 'start', 'message': '开始解析问题...'})

        # 1) 读取 Excel 表头（用于约束 LLM 产出列名）
        try:
            columns = await self._get_excel_columns(report_id, sheet)
        except Exception as e:
            await self._send('error', {'error': f'读取Excel表头失败: {str(e)}'})
            return

        await self._send('progress', {'stage': 'plan', 'message': '调用AI生成计算计划...'})

        # 2) 调用 AI 产出受限 JSON 计划
        try:
            plan = await self._generate_plan(question, columns)
        except Exception as e:
            await self._send('error', {'error': f'AI生成计划失败: {str(e)}'})
            return

        await self._send('progress', {'stage': 'execute', 'message': '执行pandas计算并回写Excel...'})

        # 3) 执行计划并写回 Excel
        try:
            result = await self._execute_plan_and_writeback(report_id, sheet, plan)
        except Exception as e:
            await self._send('error', {'error': f'执行计算失败: {str(e)}'})
            return

        # 4) 返回最新预览（tail 能看到追加行）
        try:
            preview = await self._excel_preview(report_id)
        except Exception as e:
            preview = None
            logger.warning(f"生成预览失败: {e}")

        payload = {
            'label': result.get('label'),
            'value': result.get('value'),
            'write_cell': result.get('write_cell'),
            'plan': plan,
            'preview': preview,
        }

        # 直接回给当前连接，同时也广播给订阅同 report 的其他连接
        await self._send('final', payload)
        if self.group_name:
            await self.channel_layer.group_send(self.group_name, {'type': 'broadcast_final', 'data': payload})

    async def broadcast_final(self, event):
        await self._send('final', event.get('data', {}))

    async def _send(self, msg_type: str, data: Any):
        await self.send(text_data=json.dumps({
            'type': msg_type,
            'data': data,
            'timestamp': self._get_timestamp(),
        }, ensure_ascii=False))

    def _get_timestamp(self) -> int:
        from django.utils import timezone
        return int(timezone.now().timestamp() * 1000)

    @database_sync_to_async
    def _get_report(self, report_id: int):
        from .models import MonthlyReport
        try:
            return MonthlyReport.objects.get(id=report_id, created_by=self.user)
        except MonthlyReport.DoesNotExist:
            return None

    @database_sync_to_async
    def _get_excel_columns(self, report_id: int, sheet: str):
        import os
        import openpyxl
        from .models import MonthlyReport

        report = MonthlyReport.objects.get(id=report_id, created_by=self.user)
        if not report.excel_file or not os.path.exists(report.excel_file.path):
            raise FileNotFoundError("Excel文件不存在")

        wb = openpyxl.load_workbook(report.excel_file.path, data_only=True)
        sheet_name = sheet if sheet in wb.sheetnames else (('账单明细' if '账单明细' in wb.sheetnames else wb.sheetnames[0]))
        ws = wb[sheet_name]

        max_col = ws.max_column or 0
        cols = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=1, column=c).value
            cols.append('' if v is None else str(v))
        return cols

    @database_sync_to_async
    def _generate_plan(self, question: str, columns: list[str]) -> Dict[str, Any]:
        """
        让 LLM 输出严格 JSON：{operation,label,lhs?,rhs?,column?}
        operation 白名单：mean_diff / mean / sum / count
        """
        from apps.ai_config.factory import ai_service_factory

        cols_str = '、'.join([c for c in columns if c])
        prompt = (
            "你是一个数据分析助手。你必须只输出严格的 JSON（不要代码块、不要解释）。\n"
            "你将基于一个 Excel 表的列名生成一个“受限计算计划”。\n"
            "\n"
            "可用 operation 只有：\n"
            "- mean_diff: 平均差，value = mean(lhs) - mean(rhs)\n"
            "- mean: 平均值，value = mean(column)\n"
            "- sum: 求和，value = sum(column)\n"
            "- count: 计数，value = count(column)（非空）\n"
            "\n"
            "输出 JSON schema：\n"
            "{\n"
            "  \"operation\": \"mean_diff|mean|sum|count\",\n"
            "  \"label\": \"用于写入Excel A列的中文项目名（简短）\",\n"
            "  \"lhs\": \"列名(仅mean_diff)\",\n"
            "  \"rhs\": \"列名(仅mean_diff)\",\n"
            "  \"column\": \"列名(mean/sum/count)\"\n"
            "}\n"
            "\n"
            f"可用列名：{cols_str}\n"
            f"用户问题：{question}\n"
        )

        service = ai_service_factory.get_service()
        resp = service.process_request({
            'type': 'text',
            'prompt': prompt,
            'service_type': 'monthly_excel_calc_plan',
            'user': self.user,
        })

        text = (resp or {}).get('generated_text') or ''
        text = text.strip()
        try:
            plan = json.loads(text)
        except Exception as e:
            raise ValueError(f"AI未输出可解析JSON: {str(e)}; raw={text[:200]}")

        if not isinstance(plan, dict):
            raise ValueError("AI计划必须是JSON对象")

        op = plan.get('operation')
        if op not in ['mean_diff', 'mean', 'sum', 'count']:
            raise ValueError(f"不支持的operation: {op}")

        label = (plan.get('label') or '').strip()
        if not label:
            raise ValueError("label 不能为空")

        # 基础列校验
        col_set = set([c for c in columns if c])
        if op == 'mean_diff':
            lhs = plan.get('lhs')
            rhs = plan.get('rhs')
            if not lhs or not rhs:
                raise ValueError("mean_diff 需要 lhs 和 rhs")
            if lhs not in col_set or rhs not in col_set:
                raise ValueError("lhs/rhs 必须来自可用列名")
        else:
            column = plan.get('column')
            if not column:
                raise ValueError(f"{op} 需要 column")
            if column not in col_set:
                raise ValueError("column 必须来自可用列名")

        return plan

    @database_sync_to_async
    def _execute_plan_and_writeback(self, report_id: int, sheet: str, plan: Dict[str, Any]) -> Dict[str, Any]:
        import os
        import openpyxl
        import pandas as pd
        from .models import MonthlyReport

        report = MonthlyReport.objects.get(id=report_id, created_by=self.user)
        if not report.excel_file or not os.path.exists(report.excel_file.path):
            raise FileNotFoundError("Excel文件不存在")

        wb = openpyxl.load_workbook(report.excel_file.path)
        sheet_name = sheet if sheet in wb.sheetnames else (('账单明细' if '账单明细' in wb.sheetnames else wb.sheetnames[0]))
        ws = wb[sheet_name]

        max_col = ws.max_column or 0
        if max_col <= 0:
            raise ValueError("Excel为空，无法计算")

        # 找到数据区域：从第2行开始，遇到“整行全空”即视为数据结束
        data_end_row = 1
        for r in range(2, (ws.max_row or 1) + 1):
            all_empty = True
            for c in range(1, max_col + 1):
                if ws.cell(row=r, column=c).value not in (None, ''):
                    all_empty = False
                    break
            if all_empty:
                data_end_row = r - 1
                break
        if data_end_row == 1:
            data_end_row = ws.max_row or 1

        headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        headers = [("" if h is None else str(h)).strip() for h in headers]

        records = []
        for r in range(2, data_end_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                row_vals.append(ws.cell(row=r, column=c).value)
            records.append(row_vals)

        df = pd.DataFrame(records, columns=headers)

        # 转数值列：尽量把涉及列转成 numeric
        def as_numeric(series: pd.Series) -> pd.Series:
            return pd.to_numeric(series, errors='coerce')

        op = plan['operation']
        label = plan['label']
        value: float

        if op == 'mean_diff':
            lhs = plan['lhs']
            rhs = plan['rhs']
            lhs_s = as_numeric(df[lhs])
            rhs_s = as_numeric(df[rhs])
            value = float(lhs_s.mean() - rhs_s.mean())
        elif op == 'mean':
            col = plan['column']
            s = as_numeric(df[col])
            value = float(s.mean())
        elif op == 'sum':
            col = plan['column']
            s = as_numeric(df[col])
            value = float(s.sum())
        elif op == 'count':
            col = plan['column']
            s = df[col]
            value = int(s.notna().sum())
        else:
            raise ValueError(f"不支持的operation: {op}")

        # 写回：末尾空一行
        write_row = (ws.max_row or 1) + 2
        ws.cell(row=write_row, column=1).value = label
        ws.cell(row=write_row, column=2).value = value

        wb.save(report.excel_file.path)

        return {
            'label': label,
            'value': value,
            'write_cell': f"A{write_row}:B{write_row}",
        }

    @database_sync_to_async
    def _excel_preview(self, report_id: int) -> Dict[str, Any]:
        """复用 views.py 的预览逻辑：返回 columns + head/tail"""
        import os
        import openpyxl
        from .models import MonthlyReport

        report = MonthlyReport.objects.get(id=report_id, created_by=self.user)
        if not report.excel_file or not os.path.exists(report.excel_file.path):
            raise FileNotFoundError("Excel文件不存在")

        wb = openpyxl.load_workbook(report.excel_file.path, data_only=True)
        sheet_name = '账单明细' if '账单明细' in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        columns = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=1, column=c).value
            columns.append('' if v is None else str(v))

        head_n = 20
        tail_n = 20
        data_start = 2
        data_end = max_row

        def read_rows(start_row: int, end_row: int):
            rows = []
            if start_row > end_row:
                return rows
            for r in range(start_row, end_row + 1):
                row_vals = []
                for c in range(1, max_col + 1):
                    v = ws.cell(row=r, column=c).value
                    row_vals.append('' if v is None else str(v))
                rows.append(row_vals)
            return rows

        head_end = min(data_start + head_n - 1, data_end)
        tail_start = max(data_start, data_end - tail_n + 1)

        rows_head = read_rows(data_start, head_end) if data_end >= data_start else []
        rows_tail = read_rows(tail_start, data_end) if data_end >= data_start else []

        return {
            'sheet': sheet_name,
            'columns': columns,
            'rows_head': rows_head,
            'rows_tail': rows_tail,
            'total_rows': max(0, data_end - 1),
        }


