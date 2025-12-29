"""
月度报表生成服务 - 移植自GUI项目的generate_monthly_report.py
"""
import os
import pandas as pd
import numpy as np
import io
from datetime import datetime, date
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import Q, Sum, Count, Avg
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import logging
import json
import re
from decimal import Decimal

logger = logging.getLogger(__name__)


class MonthlyReportService:
    """月度报表生成服务"""
    
    def __init__(self):
        self.output_dir = Path(settings.MEDIA_ROOT) / 'monthly_reports'
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.labor_cost_dir = Path(settings.MEDIA_ROOT) / 'labor_costs'
        self.labor_cost_dir.mkdir(parents=True, exist_ok=True)

        # 成本配置
        self.default_medicine_cost = 120.1  # 默认药水成本
        self.default_cma_cost_per_point = 60.0  # 默认CMA点位成本
        self.default_profit_rate = 0.05  # 默认分润比
    
    def generate_monthly_report(self, csv_file_path: str, config_data: Dict[str, Any]) -> Tuple[bytes, Dict[str, Any]]:
        """
        生成月度报表
        
        Args:
            csv_file_path: CSV文件路径
            config_data: 配置数据
            
        Returns:
            Tuple[bytes, Dict[str, Any]]: (Excel文件内容, 统计数据)
        """
        try:
            # 读取CSV数据
            df = self._read_csv_data(csv_file_path)
            
            # 数据预处理
            df = self._preprocess_data(df, config_data)
            
            # 计算分润比
            df = self._calculate_profit_rates(df, config_data)
            
            # 计算各种成本
            df = self._calculate_costs(df, config_data)
            
            # 生成Excel文件
            excel_content = self._generate_excel_file(df, config_data)
            
            # 生成统计数据
            summary_data = self._generate_summary_data(df)
            
            return excel_content, summary_data
            
        except Exception as e:
            logger.error(f"月度报表生成失败: {e}")
            raise e
    
    def _read_csv_data(self, csv_file_path: str) -> pd.DataFrame:
        """
        读取CSV数据 - 移植自GUI项目的CSV读取逻辑
        """
        try:
            # 尝试不同的编码格式
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(csv_file_path, encoding=encoding)
                    logger.info(f"成功使用 {encoding} 编码读取CSV文件")
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise Exception("无法读取CSV文件，尝试了所有编码格式")
            
            logger.info(f"CSV文件读取成功，共 {len(df)} 行数据")
            return df
            
        except Exception as e:
            logger.error(f"读取CSV文件失败: {e}")
            raise e

    def _normalize_csv_columns(self, df: pd.DataFrame):
        """
        归一化 CSV 列名（允许用户上传的表头命名不同）。

        - 必要列（按你要求）：客户姓名、客户电话、客户地址、成交金额、履约时间
        - 其它列均为非必要：缺失不会阻断预览/生成（但相关计算会自然为空/0）
        """
        df = df.copy()

        aliases = {
            "客户姓名": ["姓名", "客户名称", "客户名", "联系人", "收货人"],
            "客户电话": ["电话", "手机号", "手机", "联系电话", "联系方式"],
            "客户地址": ["地址", "收货地址"],
            "成交金额": ["金额", "实付", "实付价格", "实付金额", "支付金额", "付款金额", "订单金额"],
            "履约时间": ["时间", "下单时间", "订单时间", "创建时间", "付款时间", "成交时间"],
            # 非必要：用于兼容历史逻辑
            "商品类型": ["服务类型", "商品类别", "类型"],
            "备注": ["备注赠品", "赠品", "备注信息"],
        }

        for target, candidates in aliases.items():
            if target in df.columns:
                continue
            for c in candidates:
                if c in df.columns:
                    df = df.rename(columns={c: target})
                    break

        required = ["客户姓名", "客户电话", "客户地址", "成交金额", "履约时间"]
        missing_required = [c for c in required if c not in df.columns]
        return df, missing_required
    
    def _preprocess_data(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """
        数据预处理 - 移植自GUI项目的数据处理逻辑
        """
        try:
            # 先做列名归一化（不同CSV表头也能工作）
            df, _missing_required = self._normalize_csv_columns(df)

            # --- 兼容不同CSV列名 ---
            # 有些CSV没有“商品名称”，但可能有“商品类型”；预览/生成都不应因缺列直接崩溃
            if "商品名称" not in df.columns:
                if "商品类型" in df.columns:
                    df["商品名称"] = df["商品类型"]
                    logger.info("CSV缺少列[商品名称]，已使用[商品类型]作为替代用于检测订单判断")
                else:
                    df["商品名称"] = ""
                    logger.info("CSV缺少列[商品名称]与[商品类型]，将使用空值用于检测订单判断")

            # 添加检测订单标记列（缺列时已兜底）
            df["是检测订单"] = df["商品名称"].astype(str).str.contains("检测", na=False)
            
            # 按日期排序
            if "下单时间" in df.columns:
                df["下单时间"] = pd.to_datetime(df["下单时间"], errors='coerce')
                df = df.sort_values("下单时间")
            
            # 数据清洗
            df = self._clean_data(df)
            
            logger.info("数据预处理完成")
            return df
            
        except Exception as e:
            logger.error(f"数据预处理失败: {e}")
            raise e
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """数据清洗"""
        try:
            # 清理成交金额列
            if "成交金额" not in df.columns:
                # 允许预览继续；真正生成时如果缺少关键列，会在业务侧体现为金额均为0
                df["成交金额"] = 0
                logger.info("CSV缺少列[成交金额]，已默认填充为0")
            df["成交金额"] = pd.to_numeric(df["成交金额"], errors='coerce').fillna(0)
            
            # 清理数量列
            if "数量" not in df.columns:
                df["数量"] = 1
                logger.info("CSV缺少列[数量]，已默认填充为1")
            df["数量"] = pd.to_numeric(df["数量"], errors='coerce').fillna(1)
            
            # 清理CMA点位数量
            if "CMA点位数量" in df.columns:
                df["CMA点位数量"] = pd.to_numeric(df["CMA点位数量"], errors='coerce')
            
            return df
            
        except Exception as e:
            logger.error(f"数据清洗失败: {e}")
            return df
    
    def _calculate_profit_rates(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """
        计算分润比 - 移植自GUI项目的分润计算逻辑
        """
        try:
            uniform_profit_rate = config_data.get('uniform_profit_rate', False)
            profit_rate_value = config_data.get('profit_rate_value', 0.05)
            
            if uniform_profit_rate:
                # 统一分润比模式
                logger.info(f"使用统一分润比: {profit_rate_value}")
                
                def calculate_uniform_profit_rate(row):
                    # 检测订单分润比为0
                    if row.get("是检测订单", False):
                        return 0
                    
                    # 成交金额为0的订单分润比为0
                    if row.get("成交金额", 0) == 0:
                        return 0
                    
                    # 其他订单使用统一分润比
                    return profit_rate_value
                
                df["分润比"] = df.apply(calculate_uniform_profit_rate, axis=1)
            else:
                # 原有分润比逻辑
                df["分润比"] = self._calculate_original_profit_rates(df)
            
            # 计算分润金额
            df["分润金额"] = df["成交金额"] * df["分润比"]
            
            logger.info("分润比计算完成")
            return df
            
        except Exception as e:
            logger.error(f"分润比计算失败: {e}")
            raise e
    
    def _calculate_original_profit_rates(self, df: pd.DataFrame) -> pd.Series:
        """计算原有分润比逻辑"""
        # 这里实现原有的分润比计算逻辑
        # 根据商品类型、金额等因素计算不同的分润比
        
        profit_rates = []
        
        for _, row in df.iterrows():
            # 检测订单分润比为0
            if row.get("是检测订单", False):
                profit_rates.append(0)
                continue
            
            # 成交金额为0的订单分润比为0
            if row.get("成交金额", 0) == 0:
                profit_rates.append(0)
                continue
            
            # 根据商品名称或其他条件计算分润比
            product_name = str(row.get("商品名称", "")).lower()
            
            if "除醛" in product_name:
                profit_rates.append(0.08)  # 除醛产品8%
            elif "治理" in product_name:
                profit_rates.append(0.06)  # 治理产品6%
            else:
                profit_rates.append(0.05)  # 默认5%
        
        return pd.Series(profit_rates)
    
    def _calculate_costs(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """
        计算各种成本 - 移植自GUI项目的成本计算逻辑
        """
        try:
            medicine_cost_per_order = config_data.get('medicine_cost_per_order', 120.1)
            cma_cost_per_point = config_data.get('cma_cost_per_point', 60.0)
            
            # 计算CMA成本
            def calculate_cma_cost(row):
                if "CMA点位数量" in row.index and pd.notna(row["CMA点位数量"]):
                    try:
                        points = float(str(row["CMA点位数量"]).strip())
                        return points * cma_cost_per_point
                    except (ValueError, TypeError):
                        return None
                return None
            
            df["CMA成本"] = df.apply(calculate_cma_cost, axis=1)
            
            # 计算赠品成本
            df["赠品成本"] = self._calculate_gift_cost(df)
            
            # 计算备注赠品成本
            df["备注赠品成本"] = self._calculate_note_gift_cost(df)
            
            # 计算药水成本
            df["药水成本"] = df["数量"] * medicine_cost_per_order
            
            # 计算人工成本
            df["人工成本"] = self._calculate_labor_cost(df)
            
            logger.info("成本计算完成")
            return df
            
        except Exception as e:
            logger.error(f"成本计算失败: {e}")
            raise e
    
    def _calculate_gift_cost(self, df: pd.DataFrame) -> pd.Series:
        """计算赠品成本"""
        # 实现赠品成本计算逻辑
        gift_costs = []
        
        for _, row in df.iterrows():
            # 根据商品名称判断是否有赠品
            product_name = str(row.get("商品名称", "")).lower()
            
            if "赠品" in product_name or "免费" in product_name:
                # 赠品成本计算逻辑
                gift_costs.append(50.0)  # 假设赠品成本50元
            else:
                gift_costs.append(None)
        
        return pd.Series(gift_costs)
    
    def _calculate_note_gift_cost(self, df: pd.DataFrame) -> pd.Series:
        """计算备注赠品成本"""
        # 实现备注赠品成本计算逻辑
        note_gift_costs = []
        
        for _, row in df.iterrows():
            # 检查备注字段是否包含赠品信息
            note = str(row.get("备注", "")).lower()
            
            if "赠品" in note:
                # 从备注中提取赠品成本
                note_gift_costs.append(30.0)  # 假设备注赠品成本30元
            else:
                note_gift_costs.append(None)
        
        return pd.Series(note_gift_costs)
    
    def _calculate_labor_cost(self, df: pd.DataFrame) -> pd.Series:
        """计算人工成本"""
        # 实现人工成本计算逻辑
        labor_costs = []
        
        for _, row in df.iterrows():
            # 根据订单类型计算人工成本
            if row.get("是检测订单", False):
                labor_costs.append(100.0)  # 检测订单人工成本100元
            else:
                labor_costs.append(200.0)  # 治理订单人工成本200元
        
        return pd.Series(labor_costs)
    
    def _generate_excel_file(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> bytes:
        """
        生成Excel文件 - 移植自GUI项目的Excel生成逻辑
        """
        try:
            from io import BytesIO
            
            # 移除临时标记列
            export_df = df.drop(columns=["是检测订单"], errors='ignore')
            
            # 创建Excel文件
            excel_buffer = BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                export_df.to_excel(writer, sheet_name="账单明细", index=False)
                
                # 获取工作表对象
                worksheet = writer.sheets["账单明细"]
                
                # 应用样式
                self._apply_excel_styles(worksheet, export_df)
                
                # 添加统计行
                self._add_summary_rows(worksheet, export_df)
            
            excel_buffer.seek(0)
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Excel文件生成失败: {e}")
            raise e
    
    def _apply_excel_styles(self, worksheet, df: pd.DataFrame):
        """
        应用Excel样式 - 移植自GUI项目的样式设置
        """
        try:
            # 定义填充样式
            header_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")  # 浅绿色
            alt_row_fill = PatternFill(fill_type="solid", start_color="E2EFDA", end_color="E2EFDA")  # 更浅的绿色
            
            # 设置表头背景颜色
            for cell in worksheet[1]:
                cell.fill = header_fill
            
            # 设置数据行的交替背景颜色
            data_row_start = 2
            data_row_end = len(df) + 1
            
            for row_idx in range(data_row_start, data_row_end + 1):
                if (row_idx - data_row_start) % 2 == 1:  # 奇偶行交替
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = alt_row_fill
            
            # 设置列宽
            for i, column in enumerate(df.columns):
                column_width = max(
                    len(str(column)), 
                    df[column].astype(str).map(len).max() if len(df) > 0 else 10
                )
                worksheet.column_dimensions[chr(65 + i)].width = column_width * 1.2
            
            logger.info("Excel样式应用完成")
            
        except Exception as e:
            logger.warning(f"Excel样式应用失败: {e}")
    
    def _add_summary_rows(self, worksheet, df: pd.DataFrame):
        """添加统计行"""
        try:
            # 计算统计数据（允许任意列被移除：缺列则跳过对应统计）
            total_order_count = len(df)

            total_deal_amount = None
            if "成交金额" in df.columns:
                try:
                    total_deal_amount = df[df["成交金额"] > 0]["成交金额"].sum()
                except Exception:
                    total_deal_amount = df["成交金额"].sum()

            total_profit_amount = df["分润金额"].sum() if "分润金额" in df.columns else None
            total_cma_cost = df["CMA成本"].sum() if "CMA成本" in df.columns else None
            
            # 添加统计行
            summary_row = len(df) + 3
            
            worksheet.cell(row=summary_row, column=1).value = "统计汇总"

            row_cursor = summary_row + 1
            if total_deal_amount is not None:
                worksheet.cell(row=row_cursor, column=1).value = f"成交金额总计: {float(total_deal_amount):.2f}"
                row_cursor += 1
            if total_profit_amount is not None:
                worksheet.cell(row=row_cursor, column=1).value = f"分润金额总计: {float(total_profit_amount):.2f}"
                row_cursor += 1
            if total_cma_cost is not None:
                worksheet.cell(row=row_cursor, column=1).value = f"CMA成本总计: {float(total_cma_cost):.2f}"
                row_cursor += 1

            worksheet.cell(row=row_cursor, column=1).value = f"订单总数: {total_order_count}"
            
            logger.info("统计行添加完成")
            
        except Exception as e:
            logger.warning(f"统计行添加失败: {e}")
    
    def _generate_summary_data(self, df: pd.DataFrame) -> Dict[str, Any]:
        """生成统计数据"""
        try:
            # 过滤检测订单
            non_detection_df = df[~df.get("是检测订单", False)]
            
            summary = {
                'total_orders': len(df),
                'total_revenue': float(non_detection_df["成交金额"].sum()),
                'total_profit': float(df["分润金额"].sum()),
                'total_cma_cost': float(df["CMA成本"].sum()),
                'total_medicine_cost': float(df["药水成本"].sum()),
                'total_labor_cost': float(df["人工成本"].sum()),
                'average_order_value': float(non_detection_df["成交金额"].mean()) if len(non_detection_df) > 0 else 0,
                'profit_margin': 0
            }
            
            # 计算利润率
            if summary['total_revenue'] > 0:
                total_costs = summary['total_cma_cost'] + summary['total_medicine_cost'] + summary['total_labor_cost']
                summary['profit_margin'] = (summary['total_profit'] - total_costs) / summary['total_revenue'] * 100
            
            return summary
            
        except Exception as e:
            logger.error(f"统计数据生成失败: {e}")
            return {}

    def generate_monthly_report_from_db(self, year: int, month: int, user_id: int,
                                      config_data: Optional[Dict[str, Any]] = None,
                                      labor_cost_file: Optional[str] = None) -> Tuple[bytes, Dict[str, Any]]:
        """
        基于数据库订单记录生成月度报表

        Args:
            year: 年份
            month: 月份
            user_id: 用户ID
            config_data: 配置数据
            labor_cost_file: 人工成本文件路径，如果为None则自动检测

        Returns:
            Tuple[bytes, Dict[str, Any]]: (Excel文件内容, 统计数据)
        """
        try:
            # 获取指定月份的订单数据
            df = self._get_orders_from_db(year, month, user_id)

            if df.empty:
                raise ValueError(f"{year}年{month}月没有找到订单数据")

            # 数据预处理
            df = self._preprocess_db_data(df, config_data or {})

            # 计算分润比
            df = self._calculate_profit_rates_db(df, config_data or {})

            # 计算各种成本
            df = self._calculate_costs_db(df, config_data or {})

            # 处理人工成本
            df, labor_costs = self._process_labor_cost_db(df, year, month, labor_cost_file)

            # 生成Excel文件
            excel_content = self._generate_excel_file_db(df, config_data or {}, labor_costs)

            # 生成统计数据
            summary_data = self._generate_summary_data_db(df)

            return excel_content, summary_data

        except Exception as e:
            logger.error(f"基于数据库的月度报表生成失败: {e}")
            raise e

    def _get_orders_from_db(self, year: int, month: int, user_id: int) -> pd.DataFrame:
        """从数据库获取指定月份的订单数据"""
        try:
            from apps.ocr.models import CSVRecord

            # 查询指定月份的订单记录（显示所有用户的记录）
            orders = CSVRecord.objects.filter(
                履约时间__year=year,
                履约时间__month=month,
                is_active=True
            ).values(
                '客户姓名', '客户电话', '客户地址', '商品类型', '成交金额',
                '面积', '履约时间', 'CMA点位数量', '备注赠品', 'created_at'
            )

            if not orders.exists():
                return pd.DataFrame()

            # 转换为DataFrame
            df = pd.DataFrame(list(orders))

            # 数据类型转换
            df['成交金额'] = pd.to_numeric(df['成交金额'], errors='coerce').fillna(0)
            df['面积'] = pd.to_numeric(df['面积'], errors='coerce').fillna(0)
            df['CMA点位数量'] = pd.to_numeric(df['CMA点位数量'], errors='coerce').fillna(0)

            # 处理日期
            df['履约时间'] = pd.to_datetime(df['履约时间'])
            df['创建时间'] = pd.to_datetime(df['created_at'])

            logger.info(f"从数据库获取到 {len(df)} 条订单记录")
            return df

        except Exception as e:
            logger.error(f"从数据库获取订单数据失败: {e}")
            raise e

    def _preprocess_db_data(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """数据预处理 - 基于数据库数据"""
        try:
            # 确保必要的列存在
            required_columns = ['客户姓名', '客户电话', '客户地址', '商品类型', '成交金额',
                              '面积', '履约时间', 'CMA点位数量', '备注赠品']

            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''

            # 数据清洗
            df['客户姓名'] = df['客户姓名'].fillna('').astype(str)
            df['客户电话'] = df['客户电话'].fillna('').astype(str)
            df['客户地址'] = df['客户地址'].fillna('').astype(str)
            df['商品类型'] = df['商品类型'].fillna('').astype(str)
            df['备注赠品'] = df['备注赠品'].fillna('').astype(str)

            # 数值类型处理
            df['成交金额'] = pd.to_numeric(df['成交金额'], errors='coerce').fillna(0)
            df['面积'] = pd.to_numeric(df['面积'], errors='coerce').fillna(0)
            df['CMA点位数量'] = pd.to_numeric(df['CMA点位数量'], errors='coerce').fillna(0)

            # 添加序号列
            df.reset_index(drop=True, inplace=True)
            df['序号'] = df.index + 1

            logger.info(f"数据预处理完成，共 {len(df)} 条记录")
            return df

        except Exception as e:
            logger.error(f"数据预处理失败: {e}")
            raise e

    def _calculate_profit_rates_db(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """计算分润比 - 基于数据库数据，匹配GUI版本逻辑"""
        try:
            uniform_profit_rate = config_data.get('uniform_profit_rate', False)

            # 添加检测订单标记（简化版本，基于成交金额判断）
            df['是检测订单'] = df['成交金额'] <= 0

            # 按照GUI版本的逻辑计算分润比
            non_testing_count = 0
            profit_rates = []

            for idx, row in df.iterrows():
                if row['是检测订单']:
                    # 检测订单分润比为0
                    profit_rate = 0
                else:
                    # 非检测订单，更新计数
                    non_testing_count += 1

                    if uniform_profit_rate:
                        # 统一分润比模式：所有非检测订单都是0.05
                        profit_rate = 0.05
                    else:
                        # 原有逻辑：前5个非检测订单分润比为0.05，之后为0.08
                        profit_rate = 0.05 if non_testing_count <= 5 else 0.08

                profit_rates.append(profit_rate)

            # 将分润比列表添加到DataFrame
            df['分润比'] = profit_rates

            # 计算分润金额
            df['分润金额'] = df['成交金额'] * df['分润比']

            logger.info("分润比计算完成")
            return df

        except Exception as e:
            logger.error(f"分润比计算失败: {e}")
            raise e

    def _calculate_costs_db(self, df: pd.DataFrame, config_data: Dict[str, Any]) -> pd.DataFrame:
        """计算各种成本 - 基于数据库数据，匹配GUI版本逻辑"""
        try:
            # CMA成本计算
            def calculate_cma_cost(row):
                cma_points = row.get('CMA点位数量', 0)
                if pd.notna(cma_points) and str(cma_points).strip():
                    try:
                        points = float(str(cma_points).strip())
                        return points * 60  # 每个点位60元
                    except (ValueError, TypeError):
                        return None
                return None

            df['CMA成本'] = df.apply(calculate_cma_cost, axis=1)

            # 赠品成本计算（基于面积和商品类型）
            def calculate_gift_cost(row):
                import math
                service_type = str(row.get('商品类型', '')).strip()
                area_str = str(row.get('面积', '')).strip()

                if not area_str:
                    return None

                try:
                    area = float(area_str)
                except (ValueError, TypeError):
                    return None

                if area <= 0:
                    return None

                # 根据服务类型计算赠品成本
                if "母婴" in service_type:
                    gift_bottles = math.floor(area / 10)  # 除醛宝数量
                    carbon_bags = math.floor(area / 50)   # 炭包数量
                    total_cost = (gift_bottles * 10) + (carbon_bags * 15)
                    return total_cost
                elif "国标" in service_type:
                    gift_bottles = math.floor(area / 10)  # 除醛宝数量
                    total_cost = gift_bottles * 10
                    return total_cost
                else:
                    return None

            df['赠品成本'] = df.apply(calculate_gift_cost, axis=1)

            # 备注赠品成本计算
            df['备注赠品成本'] = df['备注赠品'].apply(self._parse_gift_cost)

            # 药水成本和人工成本在汇总行显示，数据行为None
            df['药水成本'] = None
            df['人工成本'] = None

            logger.info("成本计算完成")
            return df

        except Exception as e:
            logger.error(f"成本计算失败: {e}")
            raise e

    def _parse_gift_cost(self, gift_str: str) -> float:
        """解析赠品成本"""
        if not gift_str or gift_str.strip() == '':
            return 0.0

        try:
            total_cost = 0.0

            # 解析礼品信息格式：{除醛宝:15;炭包:3}
            if gift_str.startswith('{') and gift_str.endswith('}'):
                # 去掉大括号并按分号分割
                content = gift_str[1:-1]  # 去掉大括号
                items = content.split(';')
                
                for item in items:
                    if ':' in item:
                        gift_type, quantity = item.split(':', 1)
                        gift_type = gift_type.strip()
                        quantity = quantity.strip()
                        
                        try:
                            qty = int(quantity)
                            if '除醛宝' in gift_type:
                                total_cost += qty * 10  # 除醛宝10元/个
                            elif '炭包' in gift_type:
                                total_cost += qty * 15  # 炭包15元/包
                            elif '除醛机' in gift_type:
                                total_cost += qty * 0   # 除醛机0元（总部承担）
                        except ValueError:
                            logger.warning(f"无法解析礼品数量: {item}")
            else:
                # 简单文本解析
                if '除醛宝' in gift_str:
                    # 提取数量
                    import re
                    match = re.search(r'除醛宝[：:]\s*(\d+)', gift_str)
                    if match:
                        total_cost += int(match.group(1)) * 10

                if '炭包' in gift_str:
                    match = re.search(r'炭包[：:]\s*(\d+)', gift_str)
                    if match:
                        total_cost += int(match.group(1)) * 15

            return total_cost

        except Exception as e:
            logger.warning(f"解析赠品成本失败: {gift_str}, 错误: {e}")
            return 0.0

    def _process_labor_cost_db(self, df: pd.DataFrame, year: int, month: int,
                              labor_cost_file: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, float]]:
        """
        处理人工成本 - 基于数据库数据

        Args:
            df: 数据框
            year: 年份
            month: 月份
            labor_cost_file: 人工成本文件路径

        Returns:
            Tuple[pd.DataFrame, Dict[str, float]]: (更新后的数据框, 人工成本字典)
        """
        try:
            logger.info("处理人工成本")
            labor_costs = {}

            # 确定人工成本文件路径
            if not labor_cost_file:
                labor_cost_file = self.auto_detect_labor_cost_file(year, month)

            if labor_cost_file:
                labor_costs = self.parse_labor_cost_file(labor_cost_file)
                total_labor_cost = sum(labor_costs.values())
                logger.info(f"加载人工成本文件: {labor_cost_file}")
                logger.info(f"人工成本总计: {total_labor_cost}元")
            else:
                logger.warning("未找到人工成本文件，人工成本列将显示为空")

            # 添加人工成本列（初始化为None）
            df["人工成本"] = None

            # 按日期分组处理人工成本
            if labor_costs:
                # 获取每个日期的行索引范围，用于后续单元格合并
                date_groups = defaultdict(list)
                for idx, row in df.iterrows():
                    # 将履约时间转换为日期字符串
                    if pd.notna(row["履约时间"]):
                        if isinstance(row["履约时间"], str):
                            date_str = row["履约时间"].split()[0]  # 只取日期部分
                        else:
                            date_str = row["履约时间"].strftime('%Y-%m-%d')
                        date_groups[date_str].append(idx)

                # 为每个日期组的第一行设置人工成本
                for date_str, indices in date_groups.items():
                    if date_str in labor_costs:
                        # 只在该日期的第一行设置人工成本值
                        df.loc[indices[0], "人工成本"] = labor_costs[date_str]

            logger.info("人工成本处理完成")
            return df, labor_costs

        except Exception as e:
            logger.error(f"人工成本处理失败: {e}")
            # 返回原始数据框和空的人工成本字典
            df["人工成本"] = None
            return df, {}

    def parse_labor_cost_file(self, file_path: str) -> Dict[str, float]:
        """
        解析人工成本文件 - 移植自GUI项目

        参数:
        file_path: 人工成本文件路径

        返回:
        dict: {日期字符串: 共计金额}，例如 {"2025-04-02": 345, "2025-04-03": 616}
        """
        labor_costs = {}

        if not file_path or not os.path.exists(file_path):
            logger.warning(f"人工成本文件不存在: {file_path}")
            return labor_costs

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # 按行分割内容
            lines = content.strip().split('\n')
            current_date = None
            total_cost = None

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # 匹配日期格式（如：2025-04-02）
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', line)
                if date_match:
                    current_date = date_match.group(1)
                    logger.debug(f"找到日期: {current_date}")

                # 匹配共计金额（如：共计：345元）
                cost_match = re.search(r'共计[：:]\s*(\d+(?:\.\d+)?)', line)
                if cost_match:
                    total_cost = float(cost_match.group(1))

                    # 如果成功解析到日期和金额，添加到结果中
                    if current_date and total_cost is not None:
                        labor_costs[current_date] = total_cost
                        logger.debug(f"解析人工成本: {current_date} -> {total_cost}元")
                        # 重置变量，准备解析下一个日期
                        current_date = None
                        total_cost = None

        except Exception as e:
            logger.error(f"解析人工成本文件失败: {file_path}, 错误: {str(e)}")

        logger.info(f"成功解析人工成本文件，共{len(labor_costs)}条记录")
        return labor_costs

    def auto_detect_labor_cost_file(self, year: int, month: int) -> Optional[str]:
        """
        根据年月自动检测对应的人工成本文件

        参数:
        year: 年份
        month: 月份

        返回:
        str: 人工成本文件路径，如果未找到则返回None
        """
        try:
            # 构造人工成本文件名
            labor_file_name = f"{month}月人工.txt"
            labor_file_path = self.labor_cost_dir / labor_file_name

            if labor_file_path.exists():
                logger.info(f"自动检测到人工成本文件: {labor_file_path}")
                return str(labor_file_path)
            else:
                logger.warning(f"未找到对应的人工成本文件: {labor_file_path}")

                # 如果自动检测失败，尝试查找最新的人工成本文件
                labor_files = [f for f in self.labor_cost_dir.glob("*人工.txt")]
                if labor_files:
                    # 按修改时间排序，选择最新的
                    latest_file = max(labor_files, key=lambda x: x.stat().st_mtime)
                    logger.info(f"使用最新的人工成本文件: {latest_file}")
                    return str(latest_file)

        except Exception as e:
            logger.error(f"自动检测人工成本文件失败: {str(e)}")

        return None

    def _generate_excel_file_db(self, df: pd.DataFrame, config_data: Dict[str, Any],
                               labor_costs: Optional[Dict[str, float]] = None) -> bytes:
        """生成Excel文件 - 基于数据库数据，匹配GUI版本格式"""
        try:
            # 按照GUI版本的列顺序重新排列（移除检测订单标记列）
            columns_order = [
                '履约时间', '客户姓名', '客户地址', '商品类型', '面积', '成交金额',
                'CMA点位数量', '备注赠品', '分润比', '分润金额',
                'CMA成本', '赠品成本', '备注赠品成本', '药水成本', '人工成本'
            ]

            # 确保所有列都存在
            for col in columns_order:
                if col not in df.columns:
                    if col == '备注赠品成本':
                        df[col] = None  # 备注赠品成本默认为None
                    elif col == '药水成本':
                        df[col] = None  # 药水成本在汇总行显示
                    elif col == '人工成本':
                        df[col] = None  # 人工成本在汇总行显示
                    else:
                        df[col] = ''

            # 选择并重新排序列（移除检测订单标记列）
            available_columns = [col for col in columns_order if col in df.columns]
            df_output = df[available_columns].copy()

            # 格式化数据
            df_output['成交金额'] = df_output['成交金额'].round(2)
            df_output['分润比'] = df_output['分润比'].round(4)
            df_output['分润金额'] = df_output['分润金额'].round(2)

            # 处理可能为None的列
            for col in ['CMA成本', '赠品成本', '备注赠品成本']:
                if col in df_output.columns:
                    df_output[col] = df_output[col].apply(lambda x: round(x, 2) if pd.notna(x) else None)

            # 格式化日期
            df_output['履约时间'] = df_output['履约时间'].dt.strftime('%Y-%m-%d')

            # 生成Excel文件
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 使用GUI版本的sheet名称
                df_output.to_excel(writer, sheet_name='账单明细', index=False)

                # 获取工作表
                worksheet = writer.sheets['账单明细']

                # 应用GUI版本的样式
                self._apply_gui_excel_styles(worksheet, df_output, config_data, labor_costs)

            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.error(f"生成Excel文件失败: {e}")
            raise e

    def _apply_gui_excel_styles(self, worksheet, df_output, config_data, labor_costs=None):
        """应用GUI版本的Excel样式 - 完全移植自GUI项目"""
        try:
            # 定义填充样式 - 完全按照GUI版本
            header_fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")  # 浅绿色
            alt_row_fill = PatternFill(fill_type="solid", start_color="E2EFDA", end_color="E2EFDA")  # 更浅的绿色

            # 设置表头背景颜色
            for cell in worksheet[1]:
                cell.fill = header_fill

            # 设置数据行的交替背景颜色（从第二行开始）
            data_row_start = 2  # 数据开始行（Excel以1为基准，1是表头）
            data_row_end = len(df_output) + 1  # 数据结束行

            # 获取人工成本列的索引
            labor_cost_col = 0
            for idx, col_name in enumerate(df_output.columns, start=1):
                if "人工成本" in str(col_name):
                    labor_cost_col = idx
                    break

            # 设置交替行背景颜色，排除人工成本列
            for row_idx in range(data_row_start, data_row_end + 1):
                # 使奇偶行显示不同颜色：第一行保持默认，第二行开始填充
                if (row_idx - data_row_start) % 2 == 1:  # 第二行、第四行 ...
                    for col_idx in range(1, len(df_output.columns) + 1):
                        # 跳过人工成本列，让其保持白色背景
                        if col_idx == labor_cost_col and "人工成本" in df_output.columns:
                            continue  # 跳过人工成本列，不设置背景色
                        else:
                            # 其他列按正常逻辑处理
                            worksheet.cell(row=row_idx, column=col_idx).fill = alt_row_fill

            # 处理人工成本列的单元格合并 - 完全按照GUI版本逻辑
            if "人工成本" in df_output.columns and labor_costs:
                logger.info("处理人工成本列单元格合并")

                # 获取履约时间列的索引
                date_col_idx = None
                for idx, col_name in enumerate(df_output.columns, start=1):
                    if "履约时间" in str(col_name):
                        date_col_idx = idx
                        break

                if date_col_idx is None:
                    logger.warning("未找到履约时间列，跳过人工成本单元格合并")
                else:
                    # 按日期分组处理单元格合并
                    date_groups = {}  # {日期: [行号列表]}

                    # 收集每个日期对应的行号
                    for row_idx in range(2, len(df_output) + 2):  # 从第2行开始（跳过表头）
                        date_cell = worksheet.cell(row=row_idx, column=date_col_idx)
                        row_date = str(date_cell.value).split()[0] if date_cell.value else ""  # 只取日期部分

                        if row_date not in date_groups:
                            date_groups[row_date] = []
                        date_groups[row_date].append(row_idx)

                    # 对每个有多行的日期进行单元格合并
                    labor_col_letter = get_column_letter(labor_cost_col)
                    for date_str, row_list in date_groups.items():
                        if len(row_list) > 1:  # 只有当同一日期有多行时才合并
                            start_row = min(row_list)
                            end_row = max(row_list)
                            merge_range = f"{labor_col_letter}{start_row}:{labor_col_letter}{end_row}"

                            try:
                                worksheet.merge_cells(merge_range)
                                # 设置合并单元格的垂直居中对齐
                                merged_cell = worksheet.cell(row=start_row, column=labor_cost_col)
                                merged_cell.alignment = Alignment(vertical='center', horizontal='center')
                                logger.debug(f"合并日期 {date_str} 的单元格: {merge_range}")
                            except Exception as e:
                                logger.warning(f"合并单元格失败 {merge_range}: {str(e)}")

            # 处理空值显示：将None值对应的单元格清空 - 完全按照GUI版本
            logger.info("处理空值显示")
            for row_idx in range(2, len(df_output) + 2):  # 从第2行开始（跳过表头）
                for col_idx, col_name in enumerate(df_output.columns, start=1):
                    if col_name in ["CMA成本", "赠品成本", "备注赠品成本", "药水成本", "人工成本"]:
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        # 检查是否为空值（None、NaN、"None"、空字符串等）
                        if (cell_value is None or
                            (isinstance(cell_value, str) and cell_value.lower() in ['none', 'nan', '']) or
                            (isinstance(cell_value, float) and str(cell_value).lower() == 'nan')):
                            worksheet.cell(row=row_idx, column=col_idx).value = ""  # 设置为空字符串

            # 添加汇总行 - 完全按照GUI版本
            self._add_summary_rows_gui_style(worksheet, df_output)

            # 设置列宽
            for i, column in enumerate(df_output.columns):
                column_width = max(
                    len(str(column)),
                    df_output[column].astype(str).map(len).max() if len(df_output) > 0 else 10
                )
                # 转换为Excel列宽（大约字符宽度的1.2倍）
                worksheet.column_dimensions[chr(65 + i)].width = column_width * 1.2

            logger.info("Excel样式应用完成")

        except Exception as e:
            logger.warning(f"Excel样式应用失败: {e}")

    def _add_summary_rows_gui_style(self, worksheet, df_output):
        """添加汇总行 - 完全按照GUI版本"""
        try:
            logger.info("添加汇总行")

            # 计算各项总计
            total_deal_amount = df_output["成交金额"].sum()
            total_profit_amount = df_output["分润金额"].sum()

            # 计算各种成本总计
            total_cma_cost = df_output["CMA成本"].sum() if "CMA成本" in df_output.columns else 0
            total_gift_cost = df_output["赠品成本"].sum() if "赠品成本" in df_output.columns else 0
            total_note_gift_cost = df_output["备注赠品成本"].sum() if "备注赠品成本" in df_output.columns else 0
            total_medicine_cost = df_output["药水成本"].sum() if "药水成本" in df_output.columns else 0
            total_labor_cost = df_output["人工成本"].sum() if "人工成本" in df_output.columns else 0

            # 获取列索引
            column_indices = {col: idx + 1 for idx, col in enumerate(df_output.columns)}

            deal_amount_col = column_indices.get("成交金额", 6)
            profit_amount_col = column_indices.get("分润金额", 10)
            cma_cost_col = column_indices.get("CMA成本", 11)
            gift_cost_col = column_indices.get("赠品成本", 12)
            note_gift_cost_col = column_indices.get("备注赠品成本", 13)
            medicine_cost_col = column_indices.get("药水成本", 14)
            labor_cost_col = column_indices.get("人工成本", 15)

            # 添加总金额行 (空一行后再显示)
            summary_row = len(df_output) + 3  # +1是标题行，+1是数据行，+1是空行

            # 在总金额行添加标签和值
            worksheet.cell(row=summary_row, column=1, value="总金额")
            worksheet.cell(row=summary_row, column=deal_amount_col, value=total_deal_amount)
            worksheet.cell(row=summary_row, column=profit_amount_col, value=total_profit_amount)

            # 只有当各成本列存在时才添加对应的总计
            if "CMA成本" in df_output.columns:
                worksheet.cell(row=summary_row, column=cma_cost_col, value=total_cma_cost)
            if "赠品成本" in df_output.columns:
                worksheet.cell(row=summary_row, column=gift_cost_col, value=total_gift_cost)
            if "备注赠品成本" in df_output.columns:
                worksheet.cell(row=summary_row, column=note_gift_cost_col, value=total_note_gift_cost)
            if "药水成本" in df_output.columns:
                worksheet.cell(row=summary_row, column=medicine_cost_col, value=total_medicine_cost)
            if "人工成本" in df_output.columns:
                worksheet.cell(row=summary_row, column=labor_cost_col, value=total_labor_cost)

            logger.info("汇总行添加完成")

        except Exception as e:
            logger.warning(f"添加汇总行失败: {e}")

    def _generate_summary_data_db(self, df: pd.DataFrame) -> Dict[str, Any]:
        """生成统计数据 - 基于数据库数据"""
        try:
            # 计算汇总数据（匹配GUI版本）
            non_testing_orders = df[~df['是检测订单']]
            total_order_count = len(df)  # 所有订单数量（包含检测订单）
            total_medicine_cost = 120.1 * total_order_count  # 药水成本：120.1 × 订单总数量

            summary = {
                'total_orders': len(df),
                'total_revenue': float(non_testing_orders['成交金额'].sum()),  # 不含检测订单
                'total_profit_amount': float(non_testing_orders['分润金额'].sum()),  # 不含检测订单
                'total_medicine_cost': float(total_medicine_cost),
                'total_cma_cost': float(df['CMA成本'].fillna(0).sum()),
                'total_gift_cost': float(df['赠品成本'].fillna(0).sum()),
                'total_note_gift_cost': float(df['备注赠品成本'].fillna(0).sum()),
                'average_order_amount': float(non_testing_orders['成交金额'].mean()) if len(non_testing_orders) > 0 else 0,
                'average_profit_rate': float(non_testing_orders['分润比'].mean()) if len(non_testing_orders) > 0 else 0,
            }

            # 按商品类型统计
            if not df.empty and '商品类型' in df.columns:
                product_stats = df.groupby('商品类型').agg({
                    '成交金额': ['count', 'sum', 'mean'],
                    '分润金额': 'sum',
                    'CMA成本': 'sum'
                }).round(2)

                # 将多级列索引转换为字符串键
                product_stats_dict = {}
                for product_type in product_stats.index:
                    product_stats_dict[str(product_type)] = {
                        'order_count': float(product_stats.loc[product_type, ('成交金额', 'count')]),
                        'total_amount': float(product_stats.loc[product_type, ('成交金额', 'sum')]),
                        'avg_amount': float(product_stats.loc[product_type, ('成交金额', 'mean')]),
                        'total_profit': float(product_stats.loc[product_type, ('分润金额', 'sum')]),
                        'total_cma_cost': float(product_stats.loc[product_type, ('CMA成本', 'sum')] if ('CMA成本', 'sum') in product_stats.columns else 0)
                    }
                summary['product_type_stats'] = product_stats_dict
            else:
                summary['product_type_stats'] = {}

            logger.info("统计数据生成完成")
            return summary

        except Exception as e:
            logger.error(f"生成统计数据失败: {e}")
            raise e
