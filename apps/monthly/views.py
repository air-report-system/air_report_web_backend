"""
月度报表视图
"""
import os
import logging
import io
from datetime import date
from django.http import HttpResponse
from django.db import transaction
from django.utils import timezone
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema, inline_serializer
from drf_spectacular.types import OpenApiTypes
from .models import MonthlyReport, MonthlyReportConfig
from .serializers import (
    MonthlyReportSerializer,
    MonthlyReportDetailSerializer,
    MonthlyReportCreateSerializer,
    MonthlyReportGenerateSerializer,
    MonthlyReportStatsSerializer,
    MonthlyReportConfigSerializer
)
from apps.files.models import UploadedFile
from django.core.files.base import ContentFile


class MonthlyReportConfigViewSet(viewsets.ModelViewSet):
    """月度报表配置管理视图集"""
    queryset = MonthlyReportConfig.objects.all()
    serializer_class = MonthlyReportConfigSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """返回配置列表"""
        queryset = MonthlyReportConfig.objects.all()

        # 管理员可以看到所有配置，普通用户只能看到公共配置和自己的配置
        if not self.request.user.is_superuser:
            queryset = queryset.filter(
                models.Q(created_by__isnull=True) | models.Q(created_by=self.request.user)
            )

        return queryset.order_by('-is_default', 'name')

    def perform_create(self, serializer):
        """创建配置时设置创建者"""
        serializer.save(created_by=self.request.user)


class MonthlyReportViewSet(viewsets.ModelViewSet):
    """月度报表管理视图集"""
    queryset = MonthlyReport.objects.all()
    serializer_class = MonthlyReportSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """只返回当前用户的月度报表"""
        queryset = MonthlyReport.objects.filter(created_by=self.request.user)

        # 支持按月份过滤
        month = self.request.query_params.get('month')
        if month:
            try:
                year, month_num = month.split('-')
                queryset = queryset.filter(
                    report_month__year=int(year),
                    report_month__month=int(month_num)
                )
            except (ValueError, TypeError):
                pass

        # 支持按生成状态过滤
        is_generated = self.request.query_params.get('is_generated')
        if is_generated is not None:
            queryset = queryset.filter(is_generated=is_generated.lower() == 'true')

        return queryset.select_related('csv_file', 'log_file').order_by('-report_month', '-created_at')

    def get_serializer_class(self):
        """根据动作选择序列化器"""
        if self.action == 'retrieve':
            return MonthlyReportDetailSerializer
        return MonthlyReportSerializer

    def perform_create(self, serializer):
        """创建报表时设置创建者"""
        serializer.save(created_by=self.request.user)

    @extend_schema(
        summary="生成月度报表",
        description="生成指定月度报表的Excel和PDF文件",
        request=MonthlyReportGenerateSerializer,
        responses={202: {'description': '生成已开始'}}
    )
    @action(detail=True, methods=['post'])
    def generate(self, request, pk=None):
        """生成月度报表"""
        monthly_report = self.get_object()
        serializer = MonthlyReportGenerateSerializer(data=request.data)

        if serializer.is_valid():
            force_regenerate = serializer.validated_data.get('force_regenerate', False)
            generate_pdf = serializer.validated_data.get('generate_pdf', True)
            config_id = serializer.validated_data.get('config_id')

            # 检查是否已生成且不强制重新生成
            if monthly_report.is_generated and not force_regenerate:
                return Response({
                    'error': '报表已生成，如需重新生成请设置force_regenerate=true'
                }, status=status.HTTP_409_CONFLICT)

            # TODO: 调用异步月度报表生成任务
            # task = generate_monthly_report.delay(
            #     monthly_report.id,
            #     request.user.id,
            #     config_id,
            #     generate_pdf
            # )

            # 更新报表状态
            monthly_report.is_generated = False
            if config_id:
                monthly_report.config_data = {
                    **monthly_report.config_data,
                    'config_id': config_id,
                    'generate_pdf': generate_pdf
                }
            monthly_report.save()

            return Response({
                'message': '月度报表生成已开始',
                'report_id': monthly_report.id,
                # 'task_id': task.id,
                'status': 'generating'
            }, status=status.HTTP_202_ACCEPTED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="下载Excel文件",
        description="下载月度报表的Excel文件",
        responses={200: OpenApiTypes.BINARY}
    )
    @action(detail=True, methods=['get'])
    def download_excel(self, request, pk=None):
        """下载Excel文件"""
        monthly_report = self.get_object()

        if not monthly_report.excel_file or not os.path.exists(monthly_report.excel_file.path):
            return Response({
                'error': 'Excel文件不存在，请先生成报表'
            }, status=status.HTTP_404_NOT_FOUND)

        try:
            with open(monthly_report.excel_file.path, 'rb') as f:
                file_data = f.read()

            response = HttpResponse(
                file_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{monthly_report.title}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(file_data)

            return response

        except Exception as e:
            return Response({
                'error': f'下载失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @extend_schema(
        summary="Excel预览",
        description="返回已生成Excel的表头与部分行数据（head/tail），用于前端预览",
        responses={200: inline_serializer(
            name='MonthlyExcelPreviewResponse',
            fields={
                'sheet': serializers.CharField(),
                'columns': serializers.ListField(child=serializers.CharField()),
                'rows_head': serializers.ListField(child=serializers.ListField(child=serializers.CharField(allow_blank=True))),
                'rows_tail': serializers.ListField(child=serializers.ListField(child=serializers.CharField(allow_blank=True))),
                'total_rows': serializers.IntegerField(),
            }
        )}
    )
    @action(detail=True, methods=['get'], url_path='excel-preview')
    def excel_preview(self, request, pk=None):
        """Excel预览（基于已生成文件）"""
        monthly_report = self.get_object()

        if not monthly_report.excel_file or not os.path.exists(monthly_report.excel_file.path):
            return Response({
                'error': 'Excel文件不存在，请先生成报表'
            }, status=status.HTTP_404_NOT_FOUND)

        try:
            import openpyxl

            wb = openpyxl.load_workbook(monthly_report.excel_file.path, data_only=True)
            sheet_name = '账单明细' if '账单明细' in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]

            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # 读取表头（第一行）
            columns = []
            if max_col > 0 and max_row >= 1:
                for c in range(1, max_col + 1):
                    v = ws.cell(row=1, column=c).value
                    columns.append('' if v is None else str(v))

            # 读取 head/tail（不含表头）
            head_n = 20
            tail_n = 20
            data_start = 2
            data_end = max_row

            def _read_rows(start_row: int, end_row: int):
                rows = []
                if start_row > end_row:
                    return rows
                for r in range(start_row, end_row + 1):
                    row_vals = []
                    for c in range(1, max_col + 1):
                        v = ws.cell(row=r, column=c).value
                        row_vals.append('' if v is None else str(v))
                    # 如果整行为空，仍然保留（方便看到AI追加的空行/写入位置）
                    rows.append(row_vals)
                return rows

            head_end = min(data_start + head_n - 1, data_end)
            tail_start = max(data_start, data_end - tail_n + 1)

            rows_head = _read_rows(data_start, head_end)
            rows_tail = _read_rows(tail_start, data_end) if data_end >= data_start else []

            return Response({
                'sheet': sheet_name,
                'columns': columns,
                'rows_head': rows_head,
                'rows_tail': rows_tail,
                'total_rows': max(0, data_end - 1)  # 不含表头
            })

        except Exception as e:
            return Response({
                'error': f'预览失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @extend_schema(
        summary="下载PDF文件",
        description="下载月度报表的PDF文件",
        responses={200: OpenApiTypes.BINARY}
    )
    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """下载PDF文件"""
        monthly_report = self.get_object()

        if not monthly_report.pdf_file or not os.path.exists(monthly_report.pdf_file.path):
            return Response({
                'error': 'PDF文件不存在，请先生成报表'
            }, status=status.HTTP_404_NOT_FOUND)

        try:
            with open(monthly_report.pdf_file.path, 'rb') as f:
                file_data = f.read()

            response = HttpResponse(file_data, content_type='application/pdf')
            filename = f"{monthly_report.title}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Content-Length'] = len(file_data)

            return response

        except Exception as e:
            return Response({
                'error': f'下载失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class CreateMonthlyReportView(APIView):
    """创建月度报表视图"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="创建月度报表",
        description="基于CSV和日志文件创建月度报表",
        request=MonthlyReportCreateSerializer,
        responses={201: MonthlyReportSerializer}
    )
    def post(self, request):
        """创建月度报表"""
        serializer = MonthlyReportCreateSerializer(data=request.data)

        if serializer.is_valid():
            try:
                with transaction.atomic():
                    # 获取文件
                    csv_file = UploadedFile.objects.get(
                        id=serializer.validated_data['csv_file_id'],
                        created_by=request.user
                    )

                    log_file = None
                    if serializer.validated_data.get('log_file_id'):
                        log_file = UploadedFile.objects.get(
                            id=serializer.validated_data['log_file_id'],
                            created_by=request.user
                        )

                    # 准备配置数据
                    config_data = {
                        'uniform_profit_rate': serializer.validated_data.get('uniform_profit_rate', False),
                        'profit_rate_value': serializer.validated_data.get('profit_rate_value', 0.05),
                        'medicine_cost_per_order': serializer.validated_data.get('medicine_cost_per_order', 120.1),
                        'cma_cost_per_point': serializer.validated_data.get('cma_cost_per_point', 60.0),
                        'include_address_matching': serializer.validated_data.get('include_address_matching', True),
                        'exclude_recheck_records': serializer.validated_data.get('exclude_recheck_records', True),
                        'date_range_days': serializer.validated_data.get('date_range_days', 30),
                        'config_id': serializer.validated_data.get('config_id')
                    }

                    # 创建月度报表
                    monthly_report = MonthlyReport.objects.create(
                        title=serializer.validated_data['title'],
                        report_month=serializer.validated_data['report_month'],
                        csv_file=csv_file,
                        log_file=log_file,
                        config_data=config_data,
                        created_by=request.user
                    )

                # 返回创建的报表
                response_serializer = MonthlyReportSerializer(monthly_report)
                return Response(
                    response_serializer.data,
                    status=status.HTTP_201_CREATED
                )

            except UploadedFile.DoesNotExist:
                return Response({
                    'error': '文件不存在或无权限访问'
                }, status=status.HTTP_404_NOT_FOUND)
            except Exception as e:
                return Response({
                    'error': f'创建月度报表失败: {str(e)}'
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MonthlyReportStatsView(APIView):
    """月度报表统计视图"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="获取月度报表统计信息",
        description="获取当前用户的月度报表统计信息",
        responses={200: MonthlyReportStatsSerializer}
    )
    def get(self, request):
        """获取月度报表统计信息"""
        queryset = MonthlyReport.objects.filter(created_by=request.user)

        # 基本统计
        total_reports = queryset.count()
        generated_reports = queryset.filter(is_generated=True).count()
        pending_reports = total_reports - generated_reports

        # 按月份统计
        reports_by_month = {}
        for report in queryset:
            month_key = report.report_month.strftime('%Y-%m')
            if month_key not in reports_by_month:
                reports_by_month[month_key] = 0
            reports_by_month[month_key] += 1

        # 订单和收入统计
        total_orders_processed = 0
        total_revenue = 0.0

        for report in queryset.filter(is_generated=True):
            if report.summary_data:
                total_orders_processed += report.summary_data.get('total_orders', 0)
                total_revenue += report.summary_data.get('total_revenue', 0.0)

        # 平均生成时间
        generated_reports_with_time = queryset.filter(
            is_generated=True,
            created_at__isnull=False,
            generation_completed_at__isnull=False
        )

        if generated_reports_with_time.exists():
            total_generation_time = sum(
                (report.generation_completed_at - report.created_at).total_seconds()
                for report in generated_reports_with_time
            )
            average_generation_time = total_generation_time / generated_reports_with_time.count()
        else:
            average_generation_time = 0.0

        # 最近的报表
        recent_reports = queryset.select_related('csv_file', 'log_file').order_by('-created_at')[:5]

        stats_data = {
            'total_reports': total_reports,
            'generated_reports': generated_reports,
            'pending_reports': pending_reports,
            'reports_by_month': reports_by_month,
            'total_orders_processed': total_orders_processed,
            'total_revenue': total_revenue,
            'average_generation_time': average_generation_time,
            'recent_reports': recent_reports
        }

        serializer = MonthlyReportStatsSerializer(stats_data)
        return Response(serializer.data)


logger = logging.getLogger(__name__)


class LaborCostFileUploadView(APIView):
    """人工成本文件上传视图"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="上传人工成本文件",
        description="上传txt格式的人工成本文件，用于月度报表生成",
        request=inline_serializer(
            name='LaborCostFileUploadRequest',
            fields={
                'file': serializers.FileField(help_text='人工成本文件(txt格式)'),
                'month': serializers.IntegerField(help_text='月份'),
            }
        ),
        responses={200: {'description': '文件上传成功'}}
    )
    def post(self, request):
        """上传人工成本文件"""
        try:
            uploaded_file = request.FILES.get('file')
            month = request.data.get('month')

            if not uploaded_file:
                return Response(
                    {'error': '请选择要上传的文件'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if not month:
                return Response(
                    {'error': '请指定月份'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 验证文件格式
            if not uploaded_file.name.endswith('.txt'):
                return Response(
                    {'error': '人工成本文件必须是txt格式'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 保存文件到人工成本目录
            from .services import MonthlyReportService
            report_service = MonthlyReportService()

            # 构造文件名
            filename = f"{month}月人工.txt"
            file_path = report_service.labor_cost_dir / filename

            # 保存文件
            with open(file_path, 'wb') as f:
                for chunk in uploaded_file.chunks():
                    f.write(chunk)

            logger.info(f"人工成本文件已保存: {file_path}")

            # 验证文件内容
            try:
                labor_costs = report_service.parse_labor_cost_file(str(file_path))
                total_cost = sum(labor_costs.values())

                return Response({
                    'success': True,
                    'message': '人工成本文件上传成功',
                    'file_path': str(file_path),
                    'parsed_data': {
                        'total_records': len(labor_costs),
                        'total_cost': total_cost,
                        'dates': list(labor_costs.keys())
                    }
                })
            except Exception as parse_error:
                # 如果解析失败，删除文件
                if file_path.exists():
                    file_path.unlink()
                return Response(
                    {'error': f'文件格式错误，解析失败: {str(parse_error)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        except Exception as e:
            logger.error(f"人工成本文件上传失败: {e}")
            return Response(
                {'error': f'上传失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class GenerateMonthlyReportFromDBView(APIView):
    """基于数据库订单记录生成月度报表"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="基于数据库生成月度报表",
        description="从数据库订单记录生成月度报表，不依赖CSV文件",
        request=inline_serializer(
            name='GenerateFromDBRequest',
            fields={
                'year': serializers.IntegerField(help_text='年份'),
                'month': serializers.IntegerField(help_text='月份'),
                'title': serializers.CharField(help_text='报表标题', required=False),
                'config_data': serializers.JSONField(help_text='配置数据', required=False),
                'labor_cost_file': serializers.FileField(help_text='人工成本文件(txt格式)', required=False),
            }
        ),
        responses={200: {'description': '报表生成成功'}}
    )
    def post(self, request):
        """生成基于数据库的月度报表"""
        try:
            year = request.data.get('year')
            month = request.data.get('month')

            # 确保年份和月份是整数类型
            try:
                year = int(year) if year else None
                month = int(month) if month else None
            except (ValueError, TypeError):
                return Response(
                    {'error': '年份和月份必须是有效的数字'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            title = request.data.get('title', f'{year}年{month}月订单报表')
            config_data = request.data.get('config_data', {})
            labor_cost_file = request.FILES.get('labor_cost_file')

            # 处理config_data，确保它是字典类型
            if isinstance(config_data, str):
                try:
                    import json
                    config_data = json.loads(config_data)
                except (json.JSONDecodeError, ValueError):
                    config_data = {}
            elif config_data is None:
                config_data = {}

            if not year or not month:
                return Response(
                    {'error': '年份和月份不能为空'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 处理人工成本文件
            labor_cost_file_path = None
            if labor_cost_file:
                # 验证文件格式
                if not labor_cost_file.name.endswith('.txt'):
                    return Response(
                        {'error': '人工成本文件必须是txt格式'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # 保存上传的人工成本文件
                from pathlib import Path
                import tempfile

                # 创建临时文件
                with tempfile.NamedTemporaryFile(mode='w+b', suffix='.txt', delete=False) as temp_file:
                    for chunk in labor_cost_file.chunks():
                        temp_file.write(chunk)
                    labor_cost_file_path = temp_file.name

                logger.info(f"人工成本文件已保存到: {labor_cost_file_path}")

            # 创建月度报表服务
            from .services import MonthlyReportService
            report_service = MonthlyReportService()

            # 生成报表
            excel_content, summary_data = report_service.generate_monthly_report_from_db(
                year, month, request.user.id, config_data, labor_cost_file_path
            )

            # 创建报表记录（不依赖CSV文件）
            from datetime import date
            report_month = date(year, month, 1)

            # 保存Excel文件
            from django.core.files.base import ContentFile
            import time

            excel_filename = f"monthly_report_db_{year}_{month}_{int(time.time())}.xlsx"
            excel_file = ContentFile(excel_content, name=excel_filename)

            # 调试：检查summary_data的结构
            import json
            try:
                json.dumps(summary_data)
                logger.info("summary_data JSON序列化成功")
            except Exception as json_error:
                logger.error(f"summary_data JSON序列化失败: {json_error}")
                logger.error(f"summary_data内容: {summary_data}")
                # 清理summary_data，移除不能序列化的内容
                summary_data = {
                    'total_orders': summary_data.get('total_orders', 0),
                    'total_revenue': summary_data.get('total_revenue', 0),
                    'total_profit_amount': summary_data.get('total_profit_amount', 0),
                    'product_type_stats': {}
                }

            # 创建月度报表记录
            monthly_report = MonthlyReport.objects.create(
                title=title,
                report_month=report_month,
                csv_file=None,  # 基于数据库，不需要CSV文件
                config_data=config_data,
                summary_data=summary_data,
                is_generated=True,
                generation_completed_at=timezone.now(),
                created_by=request.user
            )

            # 保存Excel文件
            monthly_report.excel_file.save(excel_filename, excel_file, save=True)

            # 序列化返回结果
            serializer = MonthlyReportSerializer(monthly_report)

            # 清理临时文件
            if labor_cost_file_path and os.path.exists(labor_cost_file_path):
                try:
                    os.unlink(labor_cost_file_path)
                    logger.info(f"临时人工成本文件已清理: {labor_cost_file_path}")
                except Exception as cleanup_error:
                    logger.warning(f"清理临时文件失败: {cleanup_error}")

            return Response({
                'success': True,
                'message': '月度报表生成成功',
                'report': serializer.data
            })

        except ValueError as e:
            # 清理临时文件
            if 'labor_cost_file_path' in locals() and labor_cost_file_path and os.path.exists(labor_cost_file_path):
                try:
                    os.unlink(labor_cost_file_path)
                except Exception:
                    pass
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # 清理临时文件
            if 'labor_cost_file_path' in locals() and labor_cost_file_path and os.path.exists(labor_cost_file_path):
                try:
                    os.unlink(labor_cost_file_path)
                except Exception:
                    pass
            logger.error(f"生成月度报表失败: {e}")
            return Response(
                {'error': f'生成失败: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class PreviewMonthlyReportCSVView(APIView):
    """CSV预览：读取CSV并做基础处理，返回列与前N行，供前端列选择与预览"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="预览CSV（用于生成月度账表）",
        description="读取上传的CSV（csv_file_id），按配置做基础处理，返回表头与前N行",
        request=inline_serializer(
            name='PreviewMonthlyCSVRequest',
            fields={
                'csv_file_id': serializers.IntegerField(help_text='UploadedFile ID（spreadsheet）'),
                'uniform_profit_rate': serializers.BooleanField(required=False, default=False),
            }
        ),
        responses={200: inline_serializer(
            name='PreviewMonthlyCSVResponse',
            fields={
                'columns': serializers.ListField(child=serializers.CharField()),
                'rows_head': serializers.ListField(child=serializers.ListField(child=serializers.CharField(allow_blank=True))),
                'total_rows': serializers.IntegerField(),
            }
        )}
    )
    def post(self, request):
        try:
            csv_file_id = request.data.get('csv_file_id')
            uniform_profit_rate = request.data.get('uniform_profit_rate', False)

            if csv_file_id is None:
                return Response({'error': 'csv_file_id 不能为空'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                csv_file_id = int(csv_file_id)
            except (ValueError, TypeError):
                return Response({'error': 'csv_file_id 必须是整数'}, status=status.HTTP_400_BAD_REQUEST)

            # 获取文件（必须是当前用户上传）
            csv_file = UploadedFile.objects.get(id=csv_file_id, created_by=request.user)

            if csv_file.file_type != 'spreadsheet':
                return Response({'error': '必须是Excel/CSV文件（file_type=spreadsheet）'}, status=status.HTTP_400_BAD_REQUEST)

            if not csv_file.file or not os.path.exists(csv_file.file.path):
                return Response({'error': '文件不存在'}, status=status.HTTP_404_NOT_FOUND)

            from .services import MonthlyReportService
            svc = MonthlyReportService()

            # 预览尽量不失败：优先走“完整处理链”，失败则退化为原始CSV预览
            df_raw = svc._read_csv_data(csv_file.file.path)
            df_raw, missing_required = svc._normalize_csv_columns(df_raw)
            export_df = None
            warning = None

            if missing_required:
                warning = f"CSV缺少必要列: {', '.join(missing_required)}"

            try:
                df = svc._preprocess_data(df_raw.copy(), {'uniform_profit_rate': bool(uniform_profit_rate)})
                df = svc._calculate_profit_rates(df, {'uniform_profit_rate': bool(uniform_profit_rate)})
                df = svc._calculate_costs(df, {'uniform_profit_rate': bool(uniform_profit_rate)})
                export_df = df.drop(columns=["是检测订单"], errors='ignore')
            except Exception as e:
                extra = f"；原因为: {str(e)}"
                warning = (warning or "预览处理链失败，已退化为原始CSV预览") + extra
                logger.warning(warning)
                export_df = df_raw

            columns = [str(c) for c in export_df.columns.tolist()]

            head_n = 20
            rows_head_df = export_df.head(head_n).fillna('')
            rows_head = [[str(v) if v is not None else '' for v in row] for row in rows_head_df.values.tolist()]

            return Response({
                'columns': columns,
                'rows_head': rows_head,
                'total_rows': int(len(export_df)),
                **({'warning': warning} if warning else {}),
            })

        except UploadedFile.DoesNotExist:
            return Response({'error': '文件不存在或无权限访问'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"CSV预览失败: {e}")
            return Response({'error': f'预览失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class GenerateMonthlyReportFromCSVView(APIView):
    """从已上传CSV生成月度账表Excel并落盘，返回 MonthlyReport 信息"""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="从CSV生成月度账表（Excel）",
        description="基于已上传的CSV（csv_file_id）生成Excel，支持 selected_columns，并保存到 MonthlyReport.excel_file",
        request=inline_serializer(
            name='GenerateMonthlyFromCSVRequest',
            fields={
                'csv_file_id': serializers.IntegerField(help_text='UploadedFile ID（spreadsheet）'),
                'output_name': serializers.CharField(required=False, help_text='输出标题/文件名（不含扩展名）'),
                'uniform_profit_rate': serializers.BooleanField(required=False, default=False),
                'selected_columns': serializers.ListField(child=serializers.CharField(), required=False, help_text='保留的表头列（可空=全部）'),
            }
        ),
        responses={200: MonthlyReportSerializer}
    )
    def post(self, request):
        try:
            csv_file_id = request.data.get('csv_file_id')
            output_name = request.data.get('output_name') or ''
            uniform_profit_rate = request.data.get('uniform_profit_rate', False)
            selected_columns = request.data.get('selected_columns') or None

            if csv_file_id is None:
                return Response({'error': 'csv_file_id 不能为空'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                csv_file_id = int(csv_file_id)
            except (ValueError, TypeError):
                return Response({'error': 'csv_file_id 必须是整数'}, status=status.HTTP_400_BAD_REQUEST)

            csv_file = UploadedFile.objects.get(id=csv_file_id, created_by=request.user)
            if csv_file.file_type != 'spreadsheet':
                return Response({'error': '必须是Excel/CSV文件（file_type=spreadsheet）'}, status=status.HTTP_400_BAD_REQUEST)
            if not csv_file.file or not os.path.exists(csv_file.file.path):
                return Response({'error': '文件不存在'}, status=status.HTTP_404_NOT_FOUND)

            # 生成标题 & 报表月份（默认当前月）
            today = timezone.now().date()
            report_month = date(today.year, today.month, 1)
            title = (output_name or '').strip() or f"{today.month}月份账表"

            # 生成Excel bytes
            from .services import MonthlyReportService
            svc = MonthlyReportService()

            # 复用现有服务的处理链，最后用 _generate_excel_file 产出 bytes
            df = svc._read_csv_data(csv_file.file.path)
            df, missing_required = svc._normalize_csv_columns(df)
            if missing_required:
                return Response(
                    {
                        'error': f"CSV缺少必要列: {', '.join(missing_required)}",
                        'missing_columns': missing_required,
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            df = svc._preprocess_data(df, {'uniform_profit_rate': bool(uniform_profit_rate)})
            df = svc._calculate_profit_rates(df, {'uniform_profit_rate': bool(uniform_profit_rate)})
            df = svc._calculate_costs(df, {'uniform_profit_rate': bool(uniform_profit_rate)})

            export_df = df.drop(columns=["是检测订单"], errors='ignore')
            if selected_columns:
                # 仅保留存在的列；不存在列忽略（列选择的严格校验/提示在后续 todo 里补齐）
                selected = [c for c in selected_columns if c in export_df.columns]
                if selected:
                    export_df = export_df[selected]

            # 生成Excel bytes（借用服务的样式/汇总行）
            excel_bytes = svc._generate_excel_file(export_df, {'uniform_profit_rate': bool(uniform_profit_rate)})

            # 汇总数据（可选：先用原 df 生成，避免列裁剪影响统计；这里先保持一致）
            summary_data = svc._generate_summary_data(df)

            # 写入 MonthlyReport 记录与文件
            with transaction.atomic():
                monthly_report = MonthlyReport.objects.create(
                    title=title,
                    report_month=report_month,
                    csv_file=csv_file,
                    config_data={
                        'uniform_profit_rate': bool(uniform_profit_rate),
                        'selected_columns': selected_columns or [],
                    },
                    summary_data=summary_data,
                    is_generated=True,
                    generation_completed_at=timezone.now(),
                    created_by=request.user,
                )

                import time as _time
                excel_filename = f"monthly_report_{monthly_report.id}_{int(_time.time())}.xlsx"
                monthly_report.excel_file.save(excel_filename, ContentFile(excel_bytes), save=True)

            return Response(MonthlyReportSerializer(monthly_report).data)

        except UploadedFile.DoesNotExist:
            return Response({'error': '文件不存在或无权限访问'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"从CSV生成月度账表失败: {e}")
            return Response({'error': f'生成失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
