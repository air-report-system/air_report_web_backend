"""
月度报表生成服务 - 移植自GUI项目的generate_monthly_report.py
"""
import os
import math
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import pandas as pd
from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)


class MonthlyReportService:
    """月度报表生成服务 - 移植自GUI项目功能"""
    
    def __init__(self):
        self.output_dir = Path(settings.MEDIA_ROOT) / 'monthly_reports'
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # 成本配置
        self.cma_cost_per_point = 60  # 每个CMA点位60元
        self.gift_costs = {
            'bottle': 10,  # 除醛宝10元/个
            'carbon_bag': 15,  # 炭包15元/包
            'machine': 0  # 除醛机0元（总部承担）
        }
    
    def generate_monthly_report(self, csv_file_path: str, output_name: Optional[str] = None, 
                              uniform_profit_rate: bool = False, labor_cost_file: Optional[str] = None) -> str:
        """
        生成月度Excel报表
        
        Args:
            csv_file_path: CSV文件路径
            output_name: 输出文件名
            uniform_profit_rate: 是否统一分润比为0.05
            labor_cost_file: 人工成本文件路径
            
        Returns:
            str: 生成的Excel文件路径
        """
        try:
            # 确定输出文件名
            if not output_name:
                current_month = datetime.now().month
                output_name = f"{current_month}月份账表"
            
            # 确保输出名称不包含扩展名
            if output_name.endswith(".xlsx"):
                output_name = output_name[:-5]
            
            output_path = self.output_dir / f"{output_name}.xlsx"
            
            # 读取和验证CSV文件
            df = self._read_and_validate_csv(csv_file_path)
            
            # 处理数据
            df = self._process_data(df, uniform_profit_rate)
            
            # 计算成本
            df = self._calculate_costs(df)
            
            # 处理人工成本
            labor_cost = self._process_labor_cost(labor_cost_file, csv_file_path)
            
            # 生成Excel文件
            self._generate_excel_file(df, output_path, labor_cost)
            
            logger.info(f"月度报表生成成功: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"月度报表生成失败: {e}")
            raise e
    
    def _read_and_validate_csv(self, csv_file_path: str) -> pd.DataFrame:
        """读取和验证CSV文件"""
        # 尝试不同编码
        encodings = ["utf-8", "gbk"]
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(csv_file_path, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise ValueError("无法读取CSV文件，请检查文件编码")
        
        # 标准化列名
        df.columns = (
            df.columns.astype(str)
            .str.replace("\ufeff", "", regex=False)  # 移除BOM
            .str.strip()
        )
        
        # 验证列数（兼容7列、8列和9列）
        if len(df.columns) not in [7, 8, 9]:
            raise ValueError(f"CSV文件应包含7-9列，但检测到 {len(df.columns)} 列")
        
        # 检查必要列
        required_columns = ["履约时间", "成交金额"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"CSV文件缺少必要的列: {', '.join(missing_columns)}")
        
        # 检查成交金额空值
        empty_mask = df["成交金额"].isna() | (df["成交金额"].astype(str).str.strip() == "")
        if empty_mask.any():
            raise ValueError("CSV文件的'成交金额'列存在空值，请补充完整")
        
        # 确保CMA点位数量列存在（兼容旧格式）
        if "CMA点位数量" not in df.columns:
            logger.info("添加CMA点位数量列（旧格式CSV文件兼容）")
            df["CMA点位数量"] = ""
        
        return df
    
    def _process_data(self, df: pd.DataFrame, uniform_profit_rate: bool) -> pd.DataFrame:
        """处理数据"""
        # 清洗履约时间列
        if "履约时间" in df.columns:
            df["履约时间"] = (
                df["履约时间"].astype(str)
                .str.replace("\ufeff", "", regex=False)
                .str.strip()
            )
        
        # 转换日期格式
        df["履约时间"] = pd.to_datetime(df["履约时间"], errors="coerce")
        df["履约时间"] = df["履约时间"].dt.date
        
        # 按日期排序
        df = df.sort_values("履约时间")
        
        # 处理成交金额
        df["成交金额"] = pd.to_numeric(df["成交金额"], errors="coerce").fillna(0)
        
        # 识别检测订单
        df["是检测订单"] = df.apply(self._is_testing_order, axis=1)
        
        # 计算分润比
        df = self._calculate_profit_rates(df, uniform_profit_rate)
        
        # 计算分润金额
        df["分润金额"] = df["成交金额"] * df["分润比"]
        
        return df
    
    def _is_testing_order(self, row) -> bool:
        """判断是否为检测订单"""
        # 检查商品类型列
        service_type_col = "商品类型(国标/母婴)" if "商品类型(国标/母婴)" in row.index else "商品类型"
        service_type = str(row.get(service_type_col, "")).strip().lower()
        
        # 检测订单的关键词
        testing_keywords = ["检测", "test", "检验", "复检", "初检"]
        
        return any(keyword in service_type for keyword in testing_keywords)
    
    def _calculate_profit_rates(self, df: pd.DataFrame, uniform_profit_rate: bool) -> pd.DataFrame:
        """计算分润比"""
        non_testing_count = 0
        profit_rates = []
        
        for idx, row in df.iterrows():
            if row["是检测订单"]:
                # 检测订单分润比为0
                profit_rate = 0
            else:
                # 非检测订单，更新计数
                non_testing_count += 1
                
                if uniform_profit_rate:
                    # 统一分润比模式：所有非检测订单都是0.05
                    profit_rate = 0.05
                else:
                    # 原有逻辑：前5个非检测订单分润比为0.05，之后为0.08
                    profit_rate = 0.05 if non_testing_count <= 5 else 0.08
            
            profit_rates.append(profit_rate)
        
        df["分润比"] = profit_rates
        return df
    
    def _calculate_costs(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算各项成本"""
        # CMA成本
        df["CMA成本"] = df.apply(self._calculate_cma_cost, axis=1)
        
        # 赠品成本
        df["赠品成本"] = df.apply(self._calculate_gift_cost, axis=1)
        
        # 备注赠品成本
        df["备注赠品成本"] = df.apply(self._calculate_note_gift_cost, axis=1)
        
        # 药水成本（在汇总行显示）
        df["药水成本"] = None
        
        return df
    
    def _calculate_cma_cost(self, row) -> Optional[float]:
        """计算CMA成本"""
        if "CMA点位数量" in row.index:
            cma_points = row["CMA点位数量"]
            if pd.notna(cma_points) and str(cma_points).strip():
                try:
                    points = float(str(cma_points).strip())
                    return points * self.cma_cost_per_point
                except (ValueError, TypeError):
                    return None
        return None
    
    def _calculate_gift_cost(self, row) -> Optional[float]:
        """计算赠品成本"""
        # 获取服务类型和施工面积
        service_type_col = "商品类型(国标/母婴)" if "商品类型(国标/母婴)" in row.index else "商品类型"
        service_type = str(row.get(service_type_col, "")).strip() if pd.notna(row.get(service_type_col)) else ""
        area_str = str(row.get("面积", "")).strip() if pd.notna(row.get("面积")) else ""
        
        if not area_str:
            return None
        
        try:
            area = float(area_str)
        except (ValueError, TypeError):
            return None
        
        if area <= 0:
            return None
        
        # 根据服务类型计算赠品成本
        if "母婴" in service_type:
            # 母婴服务赠品成本计算
            gift_bottles = math.floor(area / 10)  # 除醛宝数量
            carbon_bags = math.floor(area / 50)   # 炭包数量
            
            total_cost = (gift_bottles * self.gift_costs['bottle']) + (carbon_bags * self.gift_costs['carbon_bag'])
            return total_cost
            
        elif "国标" in service_type:
            # 国标服务赠品成本计算
            gift_bottles = math.floor(area / 10)  # 除醛宝数量
            total_cost = gift_bottles * self.gift_costs['bottle']
            return total_cost
        
        return None
    
    def _calculate_note_gift_cost(self, row) -> Optional[float]:
        """计算备注赠品成本"""
        note_gift_col = "备注赠品" if "备注赠品" in row.index else None
        if not note_gift_col:
            return None
        
        note_gift_str = str(row.get(note_gift_col, "")).strip() if pd.notna(row.get(note_gift_col)) else ""
        
        if not note_gift_str:
            return None
        
        return self._parse_and_calculate_gift_cost(note_gift_str)
    
    def _parse_and_calculate_gift_cost(self, gift_str: str) -> float:
        """解析并计算赠品成本"""
        total_cost = 0.0
        
        # 解析礼品信息格式：{除醛宝:15;炭包:3}
        if gift_str.startswith('{') and gift_str.endswith('}'):
            # 去掉大括号并按分号分割
            content = gift_str[1:-1]  # 去掉大括号
            items = content.split(';')
            
            for item in items:
                if ':' in item:
                    gift_type, quantity = item.split(':', 1)
                    gift_type = gift_type.strip()
                    quantity = quantity.strip()
                    
                    try:
                        qty = int(quantity)
                        if '除醛宝' in gift_type:
                            total_cost += qty * self.gift_costs['bottle']  # 除醛宝10元/个
                        elif '炭包' in gift_type:
                            total_cost += qty * self.gift_costs['carbon_bag']  # 炭包15元/包
                        elif '除醛机' in gift_type:
                            total_cost += qty * self.gift_costs['machine']  # 除醛机0元（总部承担）
                    except ValueError:
                        logger.warning(f"无法解析礼品数量: {item}")
        else:
            # 兼容旧格式的简单文本解析
            gift_str = gift_str.lower()
            
            # 解析除醛宝
            if "除醛宝" in gift_str:
                # 尝试提取数量
                import re
                match = re.search(r'(\d+).*?除醛宝', gift_str)
                if match:
                    count = int(match.group(1))
                    total_cost += count * self.gift_costs['bottle']
            
            # 解析炭包
            if "炭包" in gift_str:
                match = re.search(r'(\d+).*?炭包', gift_str)
                if match:
                    count = int(match.group(1))
                    total_cost += count * self.gift_costs['carbon_bag']
        
        return total_cost
    
    def _process_labor_cost(self, labor_cost_file: Optional[str], csv_file_path: str) -> float:
        """处理人工成本"""
        if labor_cost_file and os.path.exists(labor_cost_file):
            try:
                with open(labor_cost_file, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    return float(content)
            except (ValueError, IOError) as e:
                logger.warning(f"读取人工成本文件失败: {e}")
        
        # 自动检测人工成本文件
        return self._auto_detect_labor_cost(csv_file_path)
    
    def _auto_detect_labor_cost(self, csv_file_path: str) -> float:
        """自动检测人工成本"""
        try:
            # 从CSV文件路径推断月份
            csv_path = Path(csv_file_path)
            labor_cost_dir = csv_path.parent / "人工成本"
            
            if not labor_cost_dir.exists():
                logger.warning("人工成本目录不存在")
                return 0.0
            
            # 查找最新的人工成本文件
            labor_files = [f for f in labor_cost_dir.glob("*人工.txt")]
            if labor_files:
                # 按修改时间排序，选择最新的
                latest_file = max(labor_files, key=lambda x: x.stat().st_mtime)
                
                with open(latest_file, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                    return float(content)
            
        except Exception as e:
            logger.warning(f"自动检测人工成本失败: {e}")
        
        return 0.0
    
    def _generate_excel_file(self, df: pd.DataFrame, output_path: Path, labor_cost: float):
        """生成Excel文件 - 移植自GUI项目的Excel生成逻辑"""
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "月度报表"

            # 设置列标题
            headers = [
                "履约时间", "客户姓名", "客户地址", "商品类型", "面积", "成交金额",
                "CMA点位数量", "备注赠品", "是检测订单", "分润比", "分润金额",
                "CMA成本", "赠品成本", "备注赠品成本", "药水成本"
            ]

            # 写入标题行
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # 写入数据行
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header, "")

                    # 处理特殊值
                    if pd.isna(value):
                        value = ""
                    elif isinstance(value, (int, float)) and value == 0:
                        value = 0

                    ws.cell(row=row_idx, column=col_idx, value=value)

            # 添加汇总行
            self._add_summary_rows(ws, df, labor_cost, len(df) + 2)

            # 设置列宽
            self._set_column_widths(ws)

            # 设置边框和格式
            self._apply_formatting(ws, len(df) + 5)  # +5 for summary rows

            # 保存文件
            wb.save(output_path)
            logger.info(f"Excel文件生成完成: {output_path}")

        except Exception as e:
            logger.error(f"Excel文件生成失败: {e}")
            raise e

    def _add_summary_rows(self, ws, df: pd.DataFrame, labor_cost: float, start_row: int):
        """添加汇总行"""
        # 计算汇总数据
        total_amount = df["成交金额"].sum()
        total_profit = df["分润金额"].sum()
        total_cma_cost = df["CMA成本"].sum()
        total_gift_cost = df["赠品成本"].sum()
        total_note_gift_cost = df["备注赠品成本"].sum()

        # 计算药水成本（每单15元）
        medicine_cost = len(df) * 15

        # 总成本
        total_cost = total_cma_cost + total_gift_cost + total_note_gift_cost + medicine_cost + labor_cost

        # 净利润
        net_profit = total_profit - total_cost

        # 添加汇总行
        summary_data = [
            ("总成交金额", total_amount),
            ("总分润金额", total_profit),
            ("总CMA成本", total_cma_cost),
            ("总赠品成本", total_gift_cost),
            ("总备注赠品成本", total_note_gift_cost),
            ("总药水成本", medicine_cost),
            ("人工成本", labor_cost),
            ("总成本", total_cost),
            ("净利润", net_profit)
        ]

        for i, (label, value) in enumerate(summary_data):
            row = start_row + i
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

    def _set_column_widths(self, ws):
        """设置列宽"""
        column_widths = {
            'A': 12,  # 履约时间
            'B': 15,  # 客户姓名
            'C': 30,  # 客户地址
            'D': 15,  # 商品类型
            'E': 10,  # 面积
            'F': 12,  # 成交金额
            'G': 12,  # CMA点位数量
            'H': 20,  # 备注赠品
            'I': 12,  # 是检测订单
            'J': 10,  # 分润比
            'K': 12,  # 分润金额
            'L': 12,  # CMA成本
            'M': 12,  # 赠品成本
            'N': 15,  # 备注赠品成本
            'O': 12   # 药水成本
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _apply_formatting(self, ws, total_rows: int):
        """应用格式化"""
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 应用边框到所有单元格
        for row in ws.iter_rows(min_row=1, max_row=total_rows, min_col=1, max_col=15):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 标题行背景色
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill


def get_monthly_report_service() -> MonthlyReportService:
    """获取月度报表服务实例"""
    return MonthlyReportService()
